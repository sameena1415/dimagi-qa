import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


class ExcelManager:

    def create_workbook(self, path):
        wb = Workbook()
        filepath = path
        wb.save(filepath)
        wb.close()

    def create_sheet(self, sheet_name, path):
        wb = load_workbook(path)
        wb.create_sheet(sheet_name)
        wb.save(path)
        wb.close()

    def sheet_new(self, path):
        wb = Workbook()
        ws1 = wb.create_sheet("Sheet_A")
        ws1.title = "Title_A"
        ws2 = wb.create_sheet("Sheet_B", 0)
        ws2.title = "Title_B"
        wb.save(filename='sample_book.xlsx')
        wb.close()

    def delete_sheet(self, sheet_name, path):
        wb = openpyxl.load_workbook(path)
        sheet_delete = wb[sheet_name]
        wb.remove(sheet_delete)
        print(wb.sheetnames)
        wb.save(path)
        wb.close()
        print("Sheet deleted")

    def update_sheet(self, sheet_name, path):
        wb = load_workbook(path)
        sheet = wb.active
        sheet.title = sheet_name
        wb.save(path)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        wb.close()
        print("Sheet updated")

    def write_data(self, sheet_name, list_data, path):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        for i in list_data:
            sheet.append(i)
        wb.save(path)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        wb.close()
        print("Excel updated")

    def get_cell_value(self, sheet_name, col_num, row_num, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        data = sheet.cell(row_num,col_num).value
        print("Cell value:", data)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        return data

    def read_excel(self, sheet_name, path):
        global data1
        global data
        data1 = ''
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        row_count = sheet.max_row
        column_count = sheet.max_column
        for i in range(1, row_count + 1):
            for j in range(1, column_count + 1):
               data = sheet.cell(row=i, column=j).value
               data1 = data1+" "+str(data)
        print("Excel data: ", data1)
        return data1

    def get_cell_data(self, sheet_name, column_name, row_num, path):
        global data, column_num
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        column_count = sheet.max_column
        for i in range(1, column_count + 1):
            if sheet.cell(row=1, column=i).value == column_name:
                column_num = i
        data = sheet.cell(row=row_num, column=column_num).value
        print("Cell data: ", data)
        return data

    def write_excel_data(self, sheet_name, row_num, column_num, value, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        sheet.cell(row=row_num, column=column_num).value = value
        wb.save(path)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        wb.close()
        print("Excel updated")

    def delete_column(self, sheet_name, column_number, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        sheet.delete_cols(column_number)
        wb.save(path)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        wb.close()
        print("Excel column deleted")

    def delete_row(self, sheet_name, row_number, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        sheet.delete_rows(row_number)
        wb.save(path)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        wb.close()
        print("Excel row deleted")

    def row_size(self, sheet_name, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        row_count = sheet.max_row
        print("Row Count: ", row_count)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        return row_count

    def col_size(self, sheet_name, path):
        wb = load_workbook(path)
        sheet = wb[sheet_name]
        column_count = sheet.max_column
        print("Column Count: ", column_count)
        sheet = wb[sheet_name]
        print("Sheet values: ",list(sheet.values))
        return column_count

    def upload_to_path(self, table_id, data_list, col_name, path):
        col = self.col_size(table_id, path)
        self.write_excel_data(table_id, 1, col + 1, col_name, path)
        row = self.row_size(table_id, path)
        self.write_data(table_id, data_list, path)

    def upload_data_typesSheet(self, type_data_list, path):
        self.write_data('types', type_data_list, path)

