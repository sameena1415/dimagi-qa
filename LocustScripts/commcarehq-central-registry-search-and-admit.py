import logging
from openpyxl import load_workbook
import random
import time

from locust import HttpUser, SequentialTaskSet, between, task, tag, events
from locust.exception import InterruptTaskSet

from user.models import UserDetails, HQUser, AppDetails
from common.args import file_path
from common.utils import load_json_data
from common.web_apps import get_app_build_info

@events.init_command_line_parser.add_listener
def _(parser):
    parser.add_argument("--domain", help="CommCare domain", required=True, env_var="COMMCARE_DOMAIN")
    parser.add_argument("--app-id", help="CommCare app id", required=True, env_var="COMMCARE_APP_ID")
    parser.add_argument("--app-config", help="Configuration of CommCare app", required=True)
    parser.add_argument("--user-details", help="Path to user details file", required=True)
    parser.add_argument("--cases-to-select", help="Path to file containing cases to use for case search", required=True)

APP_CONFIG = {}
USERS_DETAILS = []
CASES_TO_SELECT = {}

class WorkloadModelSteps(SequentialTaskSet):
    wait_time = between(5, 15)

    def on_start(self):
        self.FUNC_HOME_SCREEN = APP_CONFIG['FUNC_HOME_SCREEN']
        self.FUNC_SEARCH_AND_ADMIT_MENU = APP_CONFIG['FUNC_SEARCH_AND_ADMIT_MENU']
        self.FUNC_ADMIT_CLIENT_FORM = APP_CONFIG['FUNC_ADMIT_CLIENT_FORM']
        self.SEARCH_AND_ADMIT_INPUTS = self.FUNC_SEARCH_AND_ADMIT_MENU["inputs"]

    @tag('home_screen')
    @task
    def home_screen(self):
        self.user.hq_user.navigate_start(expected_title=self.FUNC_HOME_SCREEN['title'])

    @tag('search_and_admit_menu')
    @task
    def outgoing_referrals_menu(self):
        self.user.hq_user.navigate(
            "Open 'Search And Admit' Menu",
            data={"selections": [self.FUNC_SEARCH_AND_ADMIT_MENU['selections']]},
            expected_title=self.FUNC_SEARCH_AND_ADMIT_MENU['title']
        )

    @tag('case_search_inputs')
    @task
    def case_search_input(self):
        self.case_to_select = random.choice(list(CASES_TO_SELECT.values()))
        self.inputs = {
            "case_search_ts": self.SEARCH_AND_ADMIT_INPUTS["INPUT_CASE_SEARCH_TS"],
            "fuzzy_match_dob": self.SEARCH_AND_ADMIT_INPUTS["INPUT_FUZZY_MATCH_DOB"] 
        }

        additional_inputs = {
            "first_name": self.case_to_select["first_name"],
            "last_name": self.case_to_select["last_name"],
            "dob": self.case_to_select["dob"].strftime('%Y-%m-%d'),
            "medicaid_id": self.case_to_select["medicaid_id"],
            "reason_for_no_ssn": self.SEARCH_AND_ADMIT_INPUTS["INPUT_REASON_FOR_NO_SSN"],
            "consent_collected": self.SEARCH_AND_ADMIT_INPUTS["INPUT_CONSENT_COLLECTED"],
        }

        for key, value in additional_inputs.items():
            self.inputs[key] = value

            extra_json = {
                "query_data": {
                    "m11_results:inline": {
                        "inputs": self.inputs,
                        "execute": False,
                        "force_manual_search": True,
                        "selections": [self.FUNC_SEARCH_AND_ADMIT_MENU["selections"]]
                    }
                },
                "selections": [self.FUNC_SEARCH_AND_ADMIT_MENU["selections"]]
            }

            self.user.hq_user.navigate(
                "Input for fields in 'Search and Admit' Menu",
                data=extra_json,
                expected_title=self.FUNC_SEARCH_AND_ADMIT_MENU['title']
            )

            rng = random.randrange(1, 3)
            time.sleep(rng)

    @tag('perform_a_search_and_enter_admit_client_form')
    @task
    def perform_a_search(self):
        extra_json = {
            "query_data": {
                "m11_results:inline": {
                    "inputs": self.inputs,
                    "execute": True,
                    "force_manual_search": True,
                    "selections": [self.FUNC_SEARCH_AND_ADMIT_MENU["selections"]]
                }
            },
            "selections": [self.FUNC_SEARCH_AND_ADMIT_MENU["selections"]]
        }

        self.user.hq_user.navigate(
            "Perform a Search and enter 'Admit Client' Form",
            data=extra_json,
            expected_title=self.FUNC_ADMIT_CLIENT_FORM['title']
        )

@events.init.add_listener
def _(environment, **kw):
    try:
        app_config_path = file_path(environment.parsed_options.app_config)
        APP_CONFIG.update(load_json_data(app_config_path))
        logging.info("Loaded app config")
    except Exception as e:
        logging.error("Error loading app config: %s", e)
        raise InterruptTaskSet from e
    try:
        user_path = file_path(environment.parsed_options.user_details)
        user_data = load_json_data(user_path)["user"]
        USERS_DETAILS.extend([UserDetails(**user) for user in user_data])
        logging.info("Loaded %s users", len(USERS_DETAILS))
    except Exception as e:
        logging.error("Error loading users: %s", e)
        raise InterruptTaskSet from e
    try:
        wb = load_workbook(filename=environment.parsed_options.cases_to_select, read_only=True)
        CASES_TO_SELECT.update(_extract_data_from_sheet(wb, ["name", "first_name", "last_name", "dob", "medicaid_id"]))
    except Exception as e:
        logging.error("Error loading cases to select: %s", e)
        raise InterruptTaskSet from e

class LoginCommCareHQWithUniqueUsers(HttpUser):
    tasks = [WorkloadModelSteps]
    wait_time = between(5, 10)

    def on_start(self):
        self.domain = self.environment.parsed_options.domain
        self.host = self.environment.parsed_options.host
        self.user_detail = USERS_DETAILS.pop()

        app_details = AppDetails(
            domain=self.domain,
            app_id=self.environment.parsed_options.app_id,
        )
        self.hq_user = HQUser(self.client, self.user_detail, app_details)
        self.hq_user.login(self.domain, self.host)
        self.hq_user.app_details.build_id = self._get_build_info(self.environment.parsed_options.app_id)

    def _get_build_info(self, app_id):
        build_id = get_app_build_info(self.client, self.domain, app_id)
        if build_id:
            logging.info("Using app build: %s", build_id)
        else:
            logging.warning("No build found for app: %s", app_id)
        return build_id

def _extract_data_from_sheet(workbook, headers_of_interest):
    sheet = workbook.active
    header_col_mapping = {header: None for header in headers_of_interest}

    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        if header_value in header_col_mapping:
            header_col_mapping[header_value] = col

    if None in header_col_mapping.values():
        missing_headers = [header for header, col in header_col_mapping.items() if col is None]
        logging.error("Error: Missing headers: " + str(missing_headers))
        raise InterruptTaskSet

    name_dict = {}
    for row in range(2, sheet.max_row+1):
        name = sheet.cell(row=row, column=header_col_mapping["name"]).value
        data = {header: sheet.cell(row=row, column=col).value for header, col in header_col_mapping.items()}
        name_dict[name] = data
    return name_dict
